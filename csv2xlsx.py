#!/usr/bin/env python

import csv
import os
from pathlib import Path
# import pprint
import sys
from time import sleep
from zipfile import BadZipFile, is_zipfile, ZipFile

import config
import openpyxl as opx
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, BORDER_THIN, Side
from openpyxl.utils import get_column_letter
# from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
# from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


class ArgumentNotPassedError(Exception):
    """Error argument not passed (file csv)."""
    def __init__(self):
        self.msg = f'Не передан файл для обработки!'

    def __str__(self):
        return self.msg


class ArgumentIsFolderError(Exception):
    """Error argument is folder."""
    def __init__(self):
        self.msg = f'Передана директория, а не файл!'

    def __str__(self):
        return self.msg


class FileNotExistError(Exception):
    """Error argument file not exist."""
    def __init__(self, file_in: str):
        self.msg = f"Файл '{file_in}' не существует!"

    def __str__(self):
        return self.msg


class FileNotCSVError(Exception):
    """Error file not CSV."""
    def __init__(self):
        self.msg = f'Передан не CSV файл!'

    def __str__(self):
        return self.msg


class FileNotZIPError(Exception):
    """Error file not ZIP."""
    def __init__(self):
        self.msg = f'Неверный формат ZIP файла!'

    def __str__(self):
        return self.msg


class SaveError(Exception):
    """Error save file."""
    def __init__(self):
        self.msg = f'Ошибка сохранения файла!'

    def __str__(self):
        return self.msg


class Excel:
    def __init__(self, file_out: Path) -> None:
        self.file_out = file_out
        self.wb = None
        self.ws = None
        self.col_img_start_idx = None
        self.col_img_end_idx = None
        self.col_img_name = 'Изображения'

    def create(self) -> None:
        self.wb = opx.Workbook()
        self.ws = self.wb.worksheets[0]  # wb.active
        self.ws.title = self.file_out.stem
        self.dimensions = self.ws.column_dimensions

    def display_dimensions(self) -> None:
        max_row = self.ws.max_row
        max_col = self.ws.max_column
        print(f'Файл .csv содержит\tстолбцов:{max_col} строк:{max_row}')

    def stylization(self) -> None:
        self._stylization_header()
        self.ws.sheet_properties.outlinePr.summaryRight = False

    def _stylization_header(self) -> None:
        self.ws.auto_filter.ref = self.ws.dimensions
        self.ws.row_dimensions[1].height = config.HEADER_HEIGHT
        hdr_al, hdr_fill, thin_border, hdr_font = self._get_styles_header()

        for row_cells in self.ws.iter_rows(min_row=1, max_row=1):
            for cell in row_cells:
                cell.alignment = hdr_al
                cell.fill = hdr_fill
                cell.border = thin_border
                cell.font = hdr_font

    def set_width_column(self) -> None:
        for col in range(1, self.ws.max_column + 1):
            if col <= self.col_img_start_idx or col > self.col_img_end_idx:
                # self.dimensions[get_column_letter(col)].width = config.COL_WIDTH
                # self.dimensions[get_column_letter(col)].bestFit = True
                # self.dimensions[get_column_letter(col)].auto_size = True
                self.dimensions[get_column_letter(col)].width = config.COL_WIDTH

        # dim_holder = DimensionHolder(worksheet=self.ws)
        # for col in range(1, self.ws.max_column + 1):
            # dim_holder[get_column_letter(col)] = ColumnDimension(self.ws, min=col, max=col, width=20)
        # self.ws.column_dimensions = dim_holder
        self.set_width_col_name()

    def _get_styles_header(self) -> tuple[Alignment, PatternFill, Border, Font]:
        hdr_al = Alignment(
                          horizontal='general',
                          vertical='top',
                          text_rotation=0,
                          wrap_text=True,
                          shrink_to_fit=False,
                          indent=0
                          )
        hdr_fill = PatternFill(
                              fill_type='solid',
                              start_color='B7DEE8',
                              end_color='B7DEE8'
                              )
        thin_border = Border(
                             left=Side(border_style=BORDER_THIN, color='7F7F7F'),
                             right=Side(border_style=BORDER_THIN, color='7F7F7F'),
                             top=Side(border_style=BORDER_THIN, color='7F7F7F'),
                             bottom=Side(border_style=BORDER_THIN, color='7F7F7F')
                             )
        hdr_font = Font(
                       bold=True,
                       color='1F497D'
                       )
        return (hdr_al, hdr_fill, thin_border, hdr_font)

    def get_header_text(self) -> list[str]:
        header_text = []
        for row_cells in self.ws.iter_rows(min_row=1, max_row=1):
            for cell in row_cells:
                header_text.append(cell.value)
        return header_text

    def set_width_col_name(self) -> None:
        if config.COL_NAME_WIDTH <= 0:
            return
        header_text = self.get_header_text()
        try:
            col_name_idx = header_text.index('Название товара или услуги') + 1
        except ValueError:
            return
        self.ws.column_dimensions[get_column_letter(col_name_idx)].width = config.COL_NAME_WIDTH

    def remove_columns(self) -> None:
        if not (config.REMOVE_COLUMN and config.REMOVE_COLUMNS):
            return
        for row_cells in self.ws.iter_rows(min_row=1, max_row=1):
            for cell in reversed(row_cells):
                if cell.value in config.REMOVE_COLUMNS:
                    self.ws.delete_cols(cell.column)

    def move_columns(self) -> None:
        if not (config.ADDITIONAL_ACTIONS and config.MOVE_COLUMNS):
            return
        header_text = self.get_header_text()
        for col_in, col_to in config.MOVE_COLUMNS.items():
            if col_in not in header_text or col_to not in header_text:
                continue
            col_in_idx = header_text.index(col_in) + 1
            col_to_idx = header_text.index(col_to) + 1

            self.ws.insert_cols(col_to_idx + 1, amount=1)
            if col_in_idx > col_to_idx:
                move_col = col_to_idx - col_in_idx
                col_in_idx = col_in_idx + 1
            else:
                move_col = col_to_idx - col_in_idx + 1
            col_in_letter = get_column_letter(col_in_idx)
            # print(f'{col_in_letter}1:{col_in_letter}{self.ws.max_row}')
            self.ws.move_range(
                               f'{col_in_letter}1:{col_in_letter}{self.ws.max_row}',
                               rows=0,
                               cols=move_col
                               )
            self.ws.delete_cols(col_in_idx)

            header_text = self.get_header_text()

    def copy_data_columns(self) -> None:
        if not (config.ADDITIONAL_ACTIONS and config.COPY_DATA_COLUMNS):
            return
        header_text = self.get_header_text()
        for col_to, col_in in config.COPY_DATA_COLUMNS.items():
            if col_in not in header_text or col_to not in header_text:
                continue
            col_to_idx = header_text.index(col_to) + 1
            col_in_idx = header_text.index(col_in) + 1

            for row in range(2, self.ws.max_row + 1):
                if not self.ws.cell(row=row, column=col_to_idx).value:
                    value = self.ws.cell(row=row, column=col_in_idx).value
                    self.ws.cell(row=row, column=col_to_idx, value=value)

    def image_separation(self) -> None:
        if not (config.IMAGE_SEPARATION):
            return
        try:
            col_images_idx = self.get_col_images_idx()
        except ValueError:
            return
        images = self.get_images_data(col_images_idx)
        max_images_count = max(map(len, images))
        # pprint.pprint(images)
        # print(max_images_count)
        self.insert_image_columns(col_images_idx + 1, max_images_count - 1, name=self.col_img_name)
        self.write_images_link(col_images_idx, images)
        self.col_img_start_idx = col_images_idx
        self.col_img_end_idx = col_images_idx + max_images_count - 1
        self.hidden_images_extend_columns(self.col_img_start_idx + 1, self.col_img_end_idx)

    def get_col_images_idx(self) -> int:
        header_text = self.get_header_text()
        col_images_idx = header_text.index(self.col_img_name) + 1
        return col_images_idx

    def get_images_data(self, col_images_idx) -> list[list[str]]:
        images = []
        for row in range(2, self.ws.max_row + 1):
            value = self.ws.cell(row=row, column=col_images_idx).value
            images_product = value.strip().split()
            images.append(images_product)
        return images

    def insert_image_columns(self, idx: int, amount: int, name: str) -> None:
        self.ws.insert_cols(idx, amount)
        for col in range(amount):
            self.ws.cell(row=1, column=idx + col, value=name)

    def write_images_link(self, col_idx, images) -> None:
        style = Alignment(
                          horizontal='fill',
                          vertical='top'
                          )
        for row in range(2, self.ws.max_row + 1):
            images_product = images[row - 2]
            for col, link in enumerate(images_product):
                self.ws.cell(row=row, column=col_idx + col, value=link)
                self.ws.cell(row=row, column=col_idx + col).alignment = style

    def hidden_images_extend_columns(self, col_start: int, col_end: int) -> None:
        if not config.ADDITIONAL_ACTIONS:
            return
        # print(f'{col_start=}  {col_end=}  {get_column_letter(col_start)} {get_column_letter(col_end)}')
        # self.dimensions.group(
        #                       get_column_letter(col_start),
        #                       end=get_column_letter(col_end),
        #                       outline_level=1,
        #                       hidden=True
        #                       )
        new_dim = self.dimensions[get_column_letter(col_start)]
        new_dim.outline_level = 1
        new_dim.hidden = True
        new_dim.min = col_start
        new_dim.max = col_end
        new_dim.width = config.COL_WIDTH

    def hidden_columns(self) -> None:
        if not (config.ADDITIONAL_ACTIONS and config.HIDDEN_COLUMNS):
            return
        for row_cells in self.ws.iter_rows(min_row=1, max_row=1):
            for cell in reversed(row_cells):
                if cell.value in config.HIDDEN_COLUMNS:
                    col_letter = get_column_letter(cell.column)
                    self.dimensions.group(col_letter, col_letter, outline_level=1, hidden=True)

    def freeze_region(self) -> None:
        if not config.FREEZE_REGION:
            return
        header_text = self.get_header_text()
        try:
            freeze_idx = header_text.index(config.FREEZE_REGION) + 2
        except ValueError:
            return
        self.ws.freeze_panes = self.ws[f'{get_column_letter(freeze_idx)}2']

    def save(self) -> None:
        try:
            self.wb.save(self.file_out)
        except OSError:
            raise SaveError()


def get_transferred_argument() -> str:
    try:
        arg = sys.argv[1]
    except IndexError:
        raise ArgumentNotPassedError()  # from None
    return arg


def validate_transferred_argument(arg: str) -> Path:
    file_in = Path(arg)
    if file_in.is_dir():
        raise ArgumentIsFolderError()
    elif not file_in.exists():
        raise FileNotExistError(file_in)

    if file_in.suffix == '.zip':
        file_in = unzip_file(file_in)

    if file_in.suffix != '.csv':
        raise FileNotCSVError()
    return file_in


def unzip_file(file_zip: Path) -> Path:
    if not is_zipfile(file_zip):
        raise FileNotZIPError()

    try:
        with ZipFile(file_zip, 'r') as myzip:
            # myzip.printdir()
            namelist = myzip.namelist()[0]
            unzipped_name = myzip.extract(namelist)
    except BadZipFile:
        raise FileNotZIPError()

    # rename unzipped file
    unzipped_name = Path(unzipped_name)
    correct_unzipped_name = file_zip.parent / file_zip.stem
    os.rename(unzipped_name, correct_unzipped_name)
    # print(unzipped_name.parts)
    # print(correct_unzipped_name.parts)

    remove_file(file_zip)
    return correct_unzipped_name


def remove_file(filename) -> None:
    if config.REMOVE_FILE_IN:
        os.remove(filename)


def final_message(excel: Excel) -> None:
    print(f'Оставлено столбцов:{excel.ws.max_column}')
    print('Обработка завершена')


def convert_csv(file_in: Path, excel: Excel) -> None:
    with open(file_in, 'r', encoding='utf-16') as f:
        reader = csv.reader(f, delimiter='\t')
        for row in reader:
            excel.ws.append(row)
    excel.display_dimensions()


def main() -> None:
    arg = get_transferred_argument()
    file_in = validate_transferred_argument(arg)
    print(f"Конвертируем файл '{file_in.name}'")

    file_out = file_in.parent / f'{file_in.stem}.xlsx'
    excel = Excel(file_out)
    excel.create()
    convert_csv(file_in, excel)

    excel.remove_columns()
    excel.move_columns()
    excel.copy_data_columns()
    excel.image_separation()
    excel.stylization()
    excel.set_width_column()
    excel.hidden_columns()
    excel.freeze_region()
    excel.save()

    final_message(excel)
    remove_file(file_in)


def exit_from_program(code: int = 0, close: bool = False) -> None:
    if not close:
        input('\n---------------   END   ---------------')
    else:
        sleep(1.5)
    try:
        sys.exit(code)
    except SystemExit:
        os._exit(code)


if __name__ == '__main__':
    os.system('color 71')
    try:
        main()
    except (ArgumentNotPassedError,
            ArgumentIsFolderError,
            FileNotExistError,
            FileNotCSVError,
            SaveError,
            FileNotZIPError
            ) as e:
        print(e)
        exit_from_program(code=1, close=config.CLOSECONSOLE)
    except KeyboardInterrupt:
        print('Отмена. Скрипт остановлен.')
        exit_from_program(code=0, close=config.CLOSECONSOLE)
    except Exception as ex:
        print(ex)
        if config.EXCEPTION_TRACE:
            raise ex
        exit_from_program(code=1, close=config.CLOSECONSOLE)
